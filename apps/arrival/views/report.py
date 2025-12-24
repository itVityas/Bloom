from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.http import HttpResponse
from drf_spectacular.utils import (
    extend_schema,
    extend_schema_view,
    OpenApiResponse,
)
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import (
                        Alignment, Font
                        )

from apps.arrival.serializers.report import ListOrderSerializer
from apps.arrival.permissions import ArrivalPermission
from apps.arrival.models import Order, Content, Lot, Container
from apps.invoice.models import InvoiceContainer


@extend_schema(tags=['ReportXLSX'])
@extend_schema_view(
    post=extend_schema(
        summary='Get Orders report in xlsx',
        description='Permission: admin, arrival_reader, order_writer',
        request=ListOrderSerializer,
        responses={
            200: OpenApiResponse(description="xlsx file"),
            400: OpenApiResponse(description="Missing required parameters"),
        }
    )
)
class ReportCSVView(APIView):
    permission_classes = (IsAuthenticated, ArrivalPermission)

    def post(self, request):
        orders = request.data.get('orders_id', None)
        if not orders:
            return HttpResponse(status=status.HTTP_400_BAD_REQUEST)
        orders_id = [i['id'] for i in orders if i.get('id', None)]

        wb = Workbook()
        wb.remove(wb.active)

        orders = Order.objects.filter(id__in=orders_id)
        for order in orders:
            ws = wb.create_sheet(order.name)
            ws.cell(1, 1).value = "ДВИЖЕНИЕ КОНТЕЙНЕРОВ"
            ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(1, 1).font = Font(bold=True, size=16)
            ws.merge_cells('A1:K3')
            ws.cell(4, 1).value = "График вывоза"
            ws.cell(4, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(4, 1).font = Font(bold=True, size=16)
            ws.merge_cells('A4:K4')
            ws.append([
                "Заказ",
                "Лот",
                "Specification",
                "Контейнер",
                "Товары",
                "Кол-во коробок",
                "Дата выхода",
                "Местонахождение",
                "Доставка в Витебск",
                "Статус",
                "Примечание"],
                      )
            row = ws.row_dimensions[5]
            row.font = Font(bold=True, italic=True)
            col1 = ws.column_dimensions["A"]
            col1.alignment = Alignment(wrap_text=True)
            col2 = ws.column_dimensions["B"]
            col2.width = 10
            col2.alignment = Alignment(wrap_text=True)
            col3 = ws.column_dimensions["C"]
            col3.width = 17
            col3.alignment = Alignment(wrap_text=True)
            col4 = ws.column_dimensions["D"]
            col4.width = 15
            col4.alignment = Alignment(wrap_text=True)
            col5 = ws.column_dimensions["E"]
            col5.width = 35
            col5.alignment = Alignment(wrap_text=True)
            col6 = ws.column_dimensions["F"]
            col6.width = 20
            col6.alignment = Alignment(wrap_text=True)
            col7 = ws.column_dimensions["G"]
            col7.width = 15
            col7.alignment = Alignment(wrap_text=True)
            col8 = ws.column_dimensions["H"]
            col8.width = 25
            col8.alignment = Alignment(wrap_text=True)
            col9 = ws.column_dimensions["I"]
            col9.width = 22
            col9.alignment = Alignment(wrap_text=True)
            col10 = ws.column_dimensions["J"]
            col10.width = 20
            col10.alignment = Alignment(wrap_text=True)
            col11 = ws.column_dimensions["K"]
            col11.width = 20
            col11.alignment = Alignment(wrap_text=True)

            containers = Container.objects.filter(order=order).prefetch_related('contents')
            start_position = 6
            i = 6
            if containers:
                container_name = containers[0].name
            for container in containers:
                invoice = InvoiceContainer.objects.filter(
                    container=container).first()
                lot = Lot.objects.filter(container=container).first()
                lot_name = ''
                if lot:
                    lot_name = lot.name
                number = ''
                if invoice:
                    # contract = invoice.contract
                    number = '1/' + invoice.number.split('/')[-1]
                    number += ' от ' + str(invoice.date)
                if container_name != container.name:
                    container_name = container.name
                    if start_position < i:
                        ws.merge_cells(f'D{start_position}:D{i}')
                    start_position = i + 1
                for content in container.contents.all():
                    ws.append([
                        container.order.name,                   # A
                        lot_name,                               # B
                        number,                                 # C
                        container_name,                         # D
                        content.name,                 # E
                        content.count,                # F
                        container.exit_date,            # G
                        container.location,             # H
                        container.suppose_date,         # I
                        container.state,                # J
                        container.notice,               # K
                    ])
                    i += 1
                if len(container.contents.all()) > 0:
                    ws.merge_cells(f'A6:A{container.contents.count()+5}')
                    if i != start_position:
                        ws.merge_cells(f'D{start_position}:D{i-1}')
                if len(container.contents.all()) == 0:
                    ws.append([
                        container.order.name,                   # A
                        lot_name,                               # B
                        number,                                 # C
                        container_name,                         # D
                        '',                                     # E
                        '',                                     # F
                        container.exit_date,            # G
                        container.location,             # H
                        container.suppose_date,         # I
                        container.state,                # J
                        container.notice,               # K
                    ])
                    i += 1
                    start_position = i

        file_path = "tmp/orders.xlsx"
        wb.save(file_path)
        document = open(file_path, 'rb')
        file_name = file_path.split('/')[1]
        response = HttpResponse(
            document,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
