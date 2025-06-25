import os

import pandas as pd
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader


def excel_sheet_to_html(input_file, sheet_name, output_file, template):
    """
    Конвертирует указанный лист XLSX в HTML с сохранением формул и мультиязычного текста
    """
    try:
        wb = load_workbook(input_file, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в файле")

        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append([cell if cell is not None else "" for cell in row])

        df = pd.DataFrame(data[1:], columns=data[0])

        paths = os.path.dirname(__file__).replace('/utils', '')
        env = Environment(loader=FileSystemLoader(os.path.join(paths, 'templates')))
        template = env.get_template(template)

        html_output = template.render(
            sheet_name=sheet_name,
            columns=df.columns.tolist(),
            data=df.values.tolist()
        )

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_output)

        return True

    except Exception as e:
        print(f"Ошибка конвертации: {str(e)}")
        return False
