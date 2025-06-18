import os

from openpyxl import load_workbook


def sheet_count(file) -> int:
    if not file:
        return 0
    if not os.path.exists(file.path):
        return 0
    try:
        wb = load_workbook(file)
        return len(wb.sheetnames)
    except Exception as e:
        print(e)
    return 0
