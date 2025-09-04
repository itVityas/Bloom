from openpyxl import load_workbook

from apps.invoice.models import InvoiceContainer
from apps.arrival.models import Container


def create_invoice(traindoc) -> None:
    if not traindoc:
        return None

    containers = Container.objects.filter(lot=traindoc.lot)
    list_containers = [cont.name.lower() for cont in containers]

    wb = load_workbook(traindoc.file)
    for sheet in wb.worksheets:
        number = None
        date = None
        container = None
        check_container = False
        row_count = 0
        for row in sheet.iter_rows(values_only=True, max_row=30):
            for idx, cell in enumerate(row):
                cell_value = str(cell)

                if cell_value == '№:' or cell_value == 'Specification №':
                    number = row[idx + 1]
                elif any(keyword in cell_value for keyword in ['Date / Дата:', 'Date', 'Дата']):
                    date = row[idx + 1]
                    if date == '':
                        date = sheet.cell(row=row_count, column=idx).value
                elif cell_value.lower() in list_containers:
                    container = containers.filter(name=cell_value).first()
                    check_container = True
            row_count += 1

        if check_container and number and date:
            if not InvoiceContainer.objects.filter(number=number, date=date, container=container).exists():
                InvoiceContainer.objects.create(number=number, date=date, container=container, sheet=sheet.title)
