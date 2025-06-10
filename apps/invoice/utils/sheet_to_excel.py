from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter


def sheet_to_excel(input_file, sheet_name) -> str:
    wb_big = load_workbook(input_file)
    sheet = wb_big[sheet_name]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)

    for row in sheet.iter_rows():
        for cell in row:
            new_cell = ws.cell(
                row=cell.row,
                column=cell.column,
                value=cell.value
            )

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = \
            sheet.column_dimensions[col_letter].width

    for row in range(1, sheet.max_row + 1):
        ws.row_dimensions[row].height = \
            sheet.row_dimensions[row].height

    for merged_cell in sheet.merged_cells.ranges:
        ws.merge_cells(str(merged_cell))

    output_file = f'tmp/{sheet_name}.xlsx'
    wb.save(output_file)
    return output_file
