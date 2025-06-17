import os

from openpyxl import load_workbook


def find_sheet(invoice_number, container_name, file) -> str:
    if not file:
        return None
    if not os.path.exists(file.path):
        return None
    try:
        invoice_number = invoice_number.lower()
        container_name = invoice_number.lower()
        wb = load_workbook(file)
        for sheet in wb.worksheets:
            container_check = False
            number_check = False
            for row in sheet.iter_rows(values_only=True, max_row=30):
                for cell in row:
                    if str(cell).lower().find(container_name) != -1:
                        container_check = True
                    if str(cell).lower().find(invoice_number) != -1:
                        number_check = True
                    if container_check and number_check:
                        return sheet.title
    except Exception as e:
        print(e)
    return None
