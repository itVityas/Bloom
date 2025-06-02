from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.http import HttpResponse
from drf_spectacular.utils import (
    extend_schema,
    extend_schema_view,
    OpenApiResponse,
    OpenApiParameter,
)
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import (
                        Alignment, Font
                        )

from apps.invoice.models import Invoice, InvoiceItem
from apps.invoice.permissions import InvoicePermission


@extend_schema(tags=['ReportXLSX'])
@extend_schema_view(
    get=extend_schema(
        summary='Get Orders report in xlsx',
        description='Permission: admin, arrival_reader, order_writer',
        parameters=[
            OpenApiParameter(
                name='invoice_id',
                location=OpenApiParameter.QUERY,
                description='invoice.id',
                required=True,
                type=int,
            ),
        ],
        responses={
            200: OpenApiResponse(description="xlsx file"),
            400: OpenApiResponse(description="Missing required parameters"),
        }
    )
)
class ReportXLSXInvoice(APIView):
    permission_classes = (IsAuthenticated, InvoicePermission)

    def get(self, request):
        invoice_id = request.query_params.get('invoice_id', None)
        if not invoice_id:
            return HttpResponse('Missing required params', status=status.HTTP_400_BAD_REQUEST)
        invoice = Invoice.objects.filter(id=invoice_id).first()
        if not invoice:
            return HttpResponse('Invoice not found', status=status.HTTP_400_BAD_REQUEST)
        items = InvoiceItem.objects.filter(invoice=invoice).order_by('model')

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(str(invoice.id))
        ws.cell(1, 1).value = "Invoice-specificaton / Счет-спецификация"
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, 1).font = Font(bold=True, size=16)
        ws.merge_cells('A1:K1')
        ws['A2'] = 'Договор №'
        ws['B2'] = invoice.contract
        ws.merge_cells('B2:C2')
        ws['F2'] = '№'
        ws['G2'] = invoice.number
        ws['F2'] = 'Date/Дата:'
        ws['G2'] = invoice.date
        ws['A5'] = 'Получатель:'
        ws['B5'] = invoice.buyer
        ws.merge_cells('B5:E5')
        ws['A7'] = 'Shipper/Отправитель'
        ws.merge_cells('A7:B7')
        ws['C7'] = invoice.shipper
        ws.merge_cells('C7:E7')
        ws['A8'] = 'Seller/Продавец'
        ws.merge_cells('A8:B8')
        ws['C8'] = invoice.seller
        ws.merge_cells('C8:E10')
        ws['F8'] = 'Покупатель'
        ws['F9'] = 'Buyer'
        ws['G8'] = invoice.buyer
        ws.merge_cells('G8:K10')
        ws['A12'] = 'Terms of delivery / Условия поставки:'
        ws.merge_cells('A12:B12')
        ws['C12'] = invoice.terms
        ws['F12'] = 'Страна назначения:'
        ws['G12'] = invoice.country
        ws['A13'] = 'Container/Контейнер'
        ws.merge_cells('A13:B13')
        ws['C13'] = invoice.container.name
        ws['F13'] = 'Станция назначения:'
        ws['G13'] = invoice.station
        ws['A15'] = '№'
        ws['B15'] = 'HS Code/ Код ТН ВЭД ТС'
        col2 = ws.column_dimensions["B"]
        col2.width = 15
        ws['C15'] = 'Страна происхождения товара'
        col3 = ws.column_dimensions["C"]
        col3.width = 10
        ws['D15'] = 'Description of goods'
        col4 = ws.column_dimensions["D"]
        col4.width = 30
        ws['E15'] = 'Описание товара'
        col5 = ws.column_dimensions["E"]
        col5.width = 30
        ws['F15'] = 'Qty pcs / Кол-во шт'
        col6 = ws.column_dimensions["F"]
        col6.width = 19
        ws['G15'] = 'Qty of packages/Кол-во мест, коробов'
        col7 = ws.column_dimensions["G"]
        col7.width = 19
        ws['H15'] = 'Net weight/Масса нетто кг'
        col8 = ws.column_dimensions["H"]
        col8.width = 19
        ws['I15'] = 'Gross weight/Масса брутто кг'
        col9 = ws.column_dimensions["I"]
        col9.width = 19
        ws['J15'] = 'Unit price / Цена за шт'
        col10 = ws.column_dimensions["J"]
        col10.width = 19
        ws['K15'] = 'Amount (CNY) / Стоимость'
        col11 = ws.column_dimensions["K"]
        col11.width = 19

        q_sum = 0
        n_sum = 0
        g_sum = 0
        amount = 0
        q_sum_total = 0
        n_sum_total = 0
        g_sum_total = 0
        amount_total = 0
        model_name = ''
        line_number = 16
        count = 1
        for item in items:
            if model_name != item.model:
                if model_name != '':
                    ws[f'A{line_number}'] = 'Sumtotal / Промежуточный итог'
                    ws.merge_cells(f'A{line_number}:E{line_number}')
                    ws[f'F{line_number}'] = q_sum
                    ws[f'H{line_number}'] = n_sum
                    ws[f'I{line_number}'] = g_sum
                    ws[f'K{line_number}'] = amount
                    line_number += 1
                    q_sum = 0
                    n_sum = 0
                    g_sum = 0
                    amount = 0
                ws.cell(line_number, 1).value = item.model
                ws.merge_cells(f'A{line_number}:K{line_number}')
                line_number += 1
                model_name = item.model
            ws.append([
                count,
                item.code,
                item.country,
                item.description_en,
                item.description_ru,
                item.quantity,
                invoice.packages,
                item.net_weight,
                item.gross_weight,
                item.price_pcs,
                item.price_amount
            ])
            q_sum += item.quantity
            n_sum += item.net_weight
            g_sum += item.gross_weight
            amount += item.price_amount
            q_sum_total += item.quantity
            n_sum_total += item.net_weight
            g_sum_total += item.gross_weight
            amount_total += item.price_amount
            count += 1
            line_number += 1

        ws[f'A{line_number}'] = 'Sumtotal / Промежуточный итог'
        ws.merge_cells(f'A{line_number}:E{line_number}')
        ws[f'F{line_number}'] = q_sum
        ws[f'H{line_number}'] = n_sum
        ws[f'I{line_number}'] = g_sum
        ws[f'K{line_number}'] = amount
        line_number += 1
        ws[f'A{line_number}'] = 'Total'
        ws.merge_cells(f'A{line_number}:F{line_number}')
        ws[f'K{line_number}'] = invoice.freight_cost
        line_number += 1
        ws[f'A{line_number}'] = 'Total'
        ws.merge_cells(f'A{line_number}:E{line_number}')
        ws[f'F{line_number}'] = q_sum_total
        ws[f'H{line_number}'] = n_sum_total
        ws[f'I{line_number}'] = g_sum_total
        ws[f'K{line_number}'] = amount_total + invoice.freight_cost
        line_number += 1

        ws[f'E{line_number+5}'] = "CEO / Генеральный Директор"
        ws[f'F{line_number+6}'] = "Печать / stamp"

        file_path = f"tmp/invoice{invoice.id}.xlsx"
        wb.save(file_path)
        document = open(file_path, 'rb')
        file_name = file_path.split('/')[1]
        response = HttpResponse(
            document,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
