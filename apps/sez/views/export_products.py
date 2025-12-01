# apps/sez/views/export_products.py
from io import BytesIO
import calendar

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from apps.sez.permissions import ClearanceInvoiceItemsPermission
from apps.shtrih.models import Products


@extend_schema(tags=['Clearance Workflow'])
@extend_schema(
    summary="Export Products for ClearanceInvoice",
    description="Generate and download XLSX with products linked to the specified ClearanceInvoice.",
    responses={
        200: OpenApiResponse(
            description="XLSX file",
        )
    },
)
class ClearanceInvoiceProductsExportView(APIView):
    """
    Export Products for a given ClearanceInvoice as an XLSX file.

    Excel columns:
    1) Number
    2) Product ID
    3) Model Short Name
    4) Barcode
    5) Date
    """
    permission_classes = [IsAuthenticated, ClearanceInvoiceItemsPermission]

    def get(self, request, invoice_id: int) -> HttpResponse:
        # Fetch products related to the given invoice
        products_qs = (
            Products.objects
            .filter(cleared_id=invoice_id)
            .select_related('model__name')
            .order_by('barcode')
        )

        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Invoice_{invoice_id}_Products"

        # Write header
        headers = ["Number", "Product ID", "Model Short Name", "Barcode", "Date"]
        ws.append(headers)

        # Write data rows
        model_name = ''
        prev_module = ''
        count = 0
        summarize_lines = []
        for product in products_qs:
            product_id = product.id
            model_short = product.model.name.short_name if product.model and product.model.name else ''
            barcode = product.barcode or ''
            if model_name != product.model.name:
                ws.append([])
                model_short = model_name.short_name if model_name else ''
                model_name = product.model.name
                dict_buf = {}
                dict_buf['model_name'] = model_short
                dict_buf['count'] = count
                dict_buf['module'] = prev_module
                summarize_lines.append(dict_buf)
                count = 0

            date = ''
            if barcode:
                try:
                    module = barcode[11:12]
                    mm = int(barcode[7:9])
                    yy = int('20' + barcode[9:11])
                    _, num_days = calendar.monthrange(yy, mm)
                    date = f"{yy}.{mm:02d}.{num_days}"
                except ValueError:
                    pass

            if prev_module != module:
                ws.append([])
                dict_buf = {}
                dict_buf['model_name'] = model_short
                dict_buf['count'] = count
                dict_buf['module'] = prev_module
                summarize_lines.append(dict_buf)
                prev_module = module
                count = 0

            count += 1
            ws.append([count, product_id, model_short, barcode, date])

        dict_buf = {}
        dict_buf['model_name'] = model_short
        dict_buf['count'] = count
        dict_buf['module'] = prev_module
        summarize_lines.append(dict_buf)
        prev_module = module

        # Adjust column widths and alignment
        # Set wider columns
        column_widths = {
            1: 5,   # count
            2: 15,  # Product ID
            3: 20,  # Model Short Name
            4: 30,  # Barcode
            5: 30,  # Date
        }
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Align all cells to left
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        # Add summary lines
        ws.append([])
        ws.append(["Model Short Name", "Count", "Module"])
        for line in summarize_lines:
            if line['count'] > 0:
                ws.append([line['model_name'], line['count'], line['module']])

        # Save workbook to in-memory buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Prepare response
        filename = f"clearance_invoice_{invoice_id}_products.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
